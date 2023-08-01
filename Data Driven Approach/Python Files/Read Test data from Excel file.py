import openpyxl

wk = openpyxl.load_workbook(
    "C:/Users/sai/PycharmProjects/RobotAutomation/Data Driven Approach/Test Data/Test Data.xlsx")


def fetch_number_of_rows(sheetname):
    sh = wk[sheetname]
    return sh.max_row


def fetch_cell_data(sheetname, row, col):
    sh = wk[sheetname]
    cell = sh.cell(int(row), int(col))
    return cell.value

